from PySide6.QtWidgets import QMessageBox, QTableWidgetItem, QFileDialog
from PySide6.QtGui import QColor, QFont
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from server.client import DataClient
from pages.table_search import TableWidget
import json

class EnhancedTableWidget(TableWidget):
    def __init__(self, db_handler, editable=False, can_save_data=False):
        super().__init__(editable=editable)
        self.db_handler = db_handler
        self.can_save_data = can_save_data  # 新增属性来控制保存操作的权限

    def save_data(self):
        if not self.can_save_data:
            QMessageBox.warning(self, "保存失败", "您没有权限保存数据。")
            return
        data, merged_cells = self._collect_table_data()
        result = self.db_handler.save_all(data, merged_cells)
        if result.get("status") == "success":
            QMessageBox.information(self, "保存成功", "表格数据已保存到数据库")
        else:
            QMessageBox.warning(self, "保存失败", result.get("error", "未知错误"))

    def _collect_table_data(self):
        data = []
        merged_cells = []
        for row in range(self.rowCount()):
            row_data = {}
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    row_data[str(col)] = self._parse_cell(item)
                else:
                    row_data[str(col)] = self._empty_cell_data()
            data.append(row_data)
            merged_cells += self._collect_merged_cells_info(row)
        return data, merged_cells

    def _parse_cell(self, item):
        font = item.font()
        return {
            'text': item.text(),
            'foreground': item.foreground().color().name(),
            'background': item.background().color().name(),
            'alignment': item.textAlignment(),
            'font': {
                'bold': font.bold(),
                'size': font.pointSize()
            },
            'row_height': self.rowHeight(item.row()),
            'column_width': self.columnWidth(item.column())
        }

    def _empty_cell_data(self):
        return {
            'text': '',
            'foreground': QColor(Qt.black).name(),
            'background': QColor(Qt.white).name(),
            'alignment': int(Qt.AlignLeft | Qt.AlignVCenter),
            'font': {
                'bold': False,
                'size': 10
            },
            'row_height': 20,  # 默认高度
            'column_width': 100  # 默认宽度
        }

    def _collect_merged_cells_info(self, row):
        merged_cells = []
        for col in range(self.columnCount()):
            if self.rowSpan(row, col) > 1 or self.columnSpan(row, col) > 1:
                merged_cells.append({
                    'row': row,
                    'col': col,
                    'row_span': self.rowSpan(row, col),
                    'col_span': self.columnSpan(row, col)
                })
        return merged_cells

    def load_table_data(self):
        response = self.db_handler.get_all()
        if response.get("status") == "success":
            data = response.get("data", {})
            table_data = data.get("table_data", [])
            merged_cells = data.get("merged_cells", [])

            if table_data:
                self.populate_table(table_data, merged_cells)
            else:
                self.populate_table_with_default_data()
        else:
            QMessageBox.warning(self, "加载失败", response.get("error", "未知错误"))
            self.populate_table_with_default_data()

    def populate_table(self, table_data, merged_cells):
        self.clearContents()
        self.setRowCount(len(table_data))
        self.setColumnCount(max(len(row) for row in table_data) if table_data else 0)
        for row_idx, row_data in enumerate(table_data):
            for col_idx_str, cell_data in row_data.items():
                col_idx = int(col_idx_str)
                item = QTableWidgetItem(cell_data['text'])
                item.setForeground(QColor(cell_data['foreground']))
                item.setBackground(QColor(cell_data['background']))
                item.setTextAlignment(cell_data['alignment'])
                font = QFont()
                font.setBold(cell_data['font']['bold'])
                font.setPointSize(cell_data['font']['size'])
                item.setFont(font)
                self.setItem(row_idx, col_idx, item)
                self.setRowHeight(row_idx, cell_data['row_height'])
                self.setColumnWidth(col_idx, cell_data['column_width'])
        for cell in merged_cells:
            self.setSpan(cell['row'], cell['col'], cell['row_span'], cell['col_span'])

    def refresh_data(self):
        if self.rowCount() > 0:
            self.removeRow(self.rowCount() - 1)
        self.load_table_data()
        QMessageBox.information(self, "刷新成功", "表格数据已刷新")

    def export_to_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出为Excel", "新建表格.xlsx", "Excel Files (*.xlsx)")
        if not path:  # 如果用户没有选择文件名，则使用默认文件名
            path = "新建表格.xlsx"

        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                item = self.item(row, col)
                if item:
                    cell = ws.cell(row=row + 1, column=col + 1, value=item.text())
                    alignment = item.textAlignment()
                    align = Alignment(
                        horizontal='left' if alignment & Qt.AlignLeft else
                        'right' if alignment & Qt.AlignRight else
                        'center' if alignment & Qt.AlignHCenter else 'general',
                        vertical='top' if alignment & Qt.AlignTop else
                        'bottom' if alignment & Qt.AlignBottom else
                        'center' if alignment & Qt.AlignVCenter else 'center'
                    )
                    cell.alignment = align

                    font = item.font()
                    cell.font = Font(bold=font.bold(), size=font.pointSize())

                    fg_color = item.foreground().color().name()[1:]  # Remove '#' from color code
                    bg_color = item.background().color().name()[1:]
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                    ws.row_dimensions[row + 1].height = self.rowHeight(row) / 2
                    ws.column_dimensions[get_column_letter(col + 1)].width = self.columnWidth(col) / 10

                    cell.border = thin_border

        # 合并单元格
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                if self.rowSpan(row, col) > 1 or self.columnSpan(row, col) > 1:
                    ws.merge_cells(
                        start_row=row + 1, start_column=col + 1,
                        end_row=row + self.rowSpan(row, col),
                        end_column=col + self.columnSpan(row, col)
                    )

        wb.save(path)
        QMessageBox.information(self, "导出成功", f"表格已成功导出到 {path}")

    def populate_table_with_default_data(self):
        self.setRowCount(6)
        self.setColumnCount(8)
        for row in range(6):
            for col in range(8):
                item = QTableWidgetItem(f'({row}, {col})')
                item.setForeground(QColor(Qt.black))
                item.setBackground(QColor(Qt.white))
                self.setItem(row, col, item)
