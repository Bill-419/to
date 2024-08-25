from PySide6.QtWidgets import QMessageBox,QTableWidget, QTableWidgetItem, QFileDialog
from PySide6.QtGui import QColor, QFont
from PySide6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
from server.client import DataClient
from function.table import TableWidget
import json

class TableHandler:
    def __init__(self, table_widget, db_handler, can_save_data = False):
        self.table_widget = table_widget
        self.db_handler = db_handler
        self.can_save_data = can_save_data  # 新增属性来控制保存操作的权限

        # 原始数据表格
        self.original_table = QTableWidget(self.table_widget.get_table().rowCount(), self.table_widget.get_table().columnCount())
        self.copy_table(self.table_widget.get_table(), self.original_table)

        # 副本表格，用于显示筛选结果
        self.filtered_table_widget = QTableWidget(self.original_table.rowCount(), self.original_table.columnCount())
        self.copy_table(self.original_table, self.filtered_table_widget)


    def sync_changes_to_original(self, item):
        # 获取当前单元格的行和列
        row = item.row()
        col = item.column()

        # 更新原始表格中的对应单元格
        original_item = self.original_table.item(row, col)
        if original_item:
            original_item.setText(item.text())
        else:
            self.original_table.setItem(row, col, QTableWidgetItem(item.text()))

    def copy_table(self, source_table, target_table):
        target_table.setRowCount(source_table.rowCount())
        target_table.setColumnCount(source_table.columnCount())
        target_table.setHorizontalHeaderLabels([source_table.horizontalHeaderItem(col).text() for col in range(source_table.columnCount())])
        for row in range(source_table.rowCount()):
            for col in range(source_table.columnCount()):
                source_item = source_table.item(row, col)
                if source_item:
                    target_item = QTableWidgetItem(source_item.text())
                    target_item.setForeground(source_item.foreground())
                    target_item.setBackground(source_item.background())
                    target_table.setItem(row, col, target_item)


    def save_data(self):
        if not self.can_save_data:
            QMessageBox.warning(self.table_widget, "保存失败", "您没有权限保存数据。")
            return
        data, merged_cells = self._collect_table_data()
        result = self.db_handler.save_all(data, merged_cells)
        if result.get("status") == "success":
            QMessageBox.information(self.table_widget, "保存成功", "表格数据已保存到数据库")
        else:
            QMessageBox.warning(self.table_widget, "保存失败", result.get("error", "未知错误"))

    def _collect_table_data(self):
        table = self.table_widget.get_table()
        data = []
        merged_cells = []
        for row in range(table.rowCount()):
            row_data = {}
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    row_data[str(col)] = self._parse_cell(item)
                else:
                    row_data[str(col)] = self._empty_cell_data()
            data.append(row_data)
            merged_cells += self._collect_merged_cells_info(table, row)
        return data, merged_cells

    def _parse_cell(self, item):
        font = item.font()
        return {
            'text': item.text(),
            'foreground': item.foreground().color().name(),
            'background': item.background().color().name(),
            'alignment': item.textAlignment(),
            'font': {
                'bold': font.bold(),
                'size': font.pointSize()
            },
            'row_height': item.tableWidget().rowHeight(item.row()),
            'column_width': item.tableWidget().columnWidth(item.column())
        }

    def _empty_cell_data(self):
        return {
            'text': '',
            'foreground': QColor(Qt.black).name(),
            'background': QColor(Qt.white).name(),
            'alignment': int(Qt.AlignLeft | Qt.AlignVCenter),
            'font': {
                'bold': False,
                'size': 10
            },
            'row_height': 20,  # 默认高度
            'column_width': 100  # 默认宽度
        }

    def _collect_merged_cells_info(self, table, row):
        merged_cells = []
        for col in range(table.columnCount()):
            if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                merged_cells.append({
                    'row': row,
                    'col': col,
                    'row_span': table.rowSpan(row, col),
                    'col_span': table.columnSpan(row, col)
                })
        return merged_cells

    def load_table_data(self):
        response = self.db_handler.get_all()
        if response.get("status") == "success":
            data = response.get("data", {})
            table_data = data.get("table_data", [])
            merged_cells = data.get("merged_cells", [])

            if table_data:
                self.populate_table(table_data, merged_cells)
            else:
                self.populate_table_with_default_data()
        else:
            QMessageBox.warning(self.table_widget, "加载失败", response.get("error", "未知错误"))
            self.populate_table_with_default_data()


    def populate_table(self, table_data, merged_cells):
        table = self.table_widget.get_table()
        table.clearContents()
        table.setRowCount(len(table_data))
        table.setColumnCount(max(len(row) for row in table_data) if table_data else 0)
        for row_idx, row_data in enumerate(table_data):
            for col_idx_str, cell_data in row_data.items():
                col_idx = int(col_idx_str)
                item = QTableWidgetItem(cell_data['text'])
                item.setForeground(QColor(cell_data['foreground']))
                item.setBackground(QColor(cell_data['background']))
                item.setTextAlignment(cell_data['alignment'])
                font = QFont()
                font.setBold(cell_data['font']['bold'])
                font.setPointSize(cell_data['font']['size'])
                item.setFont(font)
                table.setItem(row_idx, col_idx, item)
                table.setRowHeight(row_idx, cell_data['row_height'])
                table.setColumnWidth(col_idx, cell_data['column_width'])
        for cell in merged_cells:
            table.setSpan(cell['row'], cell['col'], cell['row_span'], cell['col_span'])

    def refresh_data(self):
        table = self.table_widget.get_table()
        if table.rowCount() > 0:
            table.removeRow(table.rowCount() - 1)
        self.load_table_data()
        QMessageBox.information(self.table_widget, "刷新成功", "表格数据已刷新")

    def export_to_excel(self):
        table = self.table_widget.get_table()
        path, _ = QFileDialog.getSaveFileName(self.table_widget, "导出为Excel", "新建表格.xlsx", "Excel Files (*.xlsx)")
        if not path:  # 如果用户没有选择文件名，则使用默认文件名
            path = "新建表格.xlsx"

        wb = Workbook()
        ws = wb.active

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    cell = ws.cell(row=row + 1, column=col + 1, value=item.text())
                    alignment = item.textAlignment()
                    align = Alignment(
                        horizontal='left' if alignment & Qt.AlignLeft else
                        'right' if alignment & Qt.AlignRight else
                        'center' if alignment & Qt.AlignHCenter else 'general',
                        vertical='top' if alignment & Qt.AlignTop else
                        'bottom' if alignment & Qt.AlignBottom else
                        'center' if alignment & Qt.AlignVCenter else 'center'
                    )
                    cell.alignment = align

                    font = item.font()
                    cell.font = Font(bold=font.bold(), size=font.pointSize())

                    fg_color = item.foreground().color().name()[1:]  # Remove '#' from color code
                    bg_color = item.background().color().name()[1:]
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

                    ws.row_dimensions[row + 1].height = table.rowHeight(row) / 2
                    ws.column_dimensions[get_column_letter(col + 1)].width = table.columnWidth(col) / 10

                    cell.border = thin_border

        # 合并单元格
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                if table.rowSpan(row, col) > 1 or table.columnSpan(row, col) > 1:
                    ws.merge_cells(
                        start_row=row + 1, start_column=col + 1,
                        end_row=row + table.rowSpan(row, col),
                        end_column=col + table.columnSpan(row, col)
                    )

        wb.save(path)
        QMessageBox.information(self.table_widget, "导出成功", f"表格已成功导出到 {path}")

    def populate_table_with_default_data(self):
        table = self.table_widget.get_table()
        table.setRowCount(2)
        table.setColumnCount(8)
        for row in range(2):
            for col in range(8):
                item = QTableWidgetItem(f'({row}, {col})')
                item.setForeground(QColor(Qt.black))
                item.setBackground(QColor(Qt.white))
                table.setItem(row, col, item)